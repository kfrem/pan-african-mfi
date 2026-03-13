"""
Data Import Validator — Pan-African Microfinance SaaS
Validates CSV/Excel uploads before importing into the system.
Supports: clients, loans, repayment history, chart of accounts, opening balances.
"""
import csv
import io
import re
import logging
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from django.utils import timezone

from apps.tenants.models import Tenant
from apps.tenants.country_pack_engine import CountryPackEngine
from apps.clients.models import Client
from apps.onboarding.models import ImportJob

logger = logging.getLogger(__name__)


class ImportValidator:
    """
    Validate and preview import data before committing to database.
    Flow: Upload → Parse → Validate → Preview → Approve → Import
    """

    def __init__(self, tenant_id: str, import_type: str):
        self.tenant = Tenant.objects.get(id=tenant_id)
        self.engine = CountryPackEngine.for_tenant(tenant_id)
        self.import_type = import_type
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []

    def parse_file(self, file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
        """Parse CSV or Excel file into list of row dicts."""
        if filename.endswith('.csv'):
            return self._parse_csv(file_bytes)
        elif filename.endswith(('.xlsx', '.xls')):
            return self._parse_excel(file_bytes)
        else:
            raise ValueError(f'Unsupported file type: {filename}. Use .csv or .xlsx')

    def _parse_csv(self, file_bytes: bytes) -> List[Dict]:
        # Try UTF-8, fall back to latin-1
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1']:
            try:
                text = file_bytes.decode(encoding)
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
                if rows:
                    return rows
            except (UnicodeDecodeError, csv.Error):
                continue
        raise ValueError('Could not parse CSV file. Check encoding.')

    def _parse_excel(self, file_bytes: bytes) -> List[Dict]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [str(cell.value or '').strip() for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            if any(v is not None for v in row_dict.values()):
                rows.append(row_dict)
        return rows

    def validate(self, rows: List[Dict]) -> Dict:
        """Run full validation on parsed rows. Returns validation report."""
        self.errors = []
        self.warnings = []

        validator = VALIDATORS.get(self.import_type)
        if not validator:
            raise ValueError(f'No validator for import type: {self.import_type}')

        # Check required columns
        if rows:
            required = validator['required_columns']
            actual = set(rows[0].keys())
            missing = set(required) - actual
            if missing:
                self.errors.append({
                    'row': 0, 'field': 'headers',
                    'error': f'Missing required columns: {", ".join(missing)}'
                })
                return self._report(len(rows))

        # Validate each row
        for idx, row in enumerate(rows, start=2):  # Row 2 is first data row
            validator['validate_row'](self, idx, row)

        return self._report(len(rows))

    def _report(self, total_rows: int) -> Dict:
        error_rows = len(set(e['row'] for e in self.errors))
        warning_rows = len(set(w['row'] for w in self.warnings))
        return {
            'total_rows': total_rows,
            'valid_rows': total_rows - error_rows,
            'error_rows': error_rows,
            'warning_rows': warning_rows,
            'errors': self.errors[:100],  # Cap at 100 for display
            'warnings': self.warnings[:100],
            'can_import': error_rows == 0,
        }

    # ─── Client Validation ───

    def _validate_client_row(self, row_num: int, row: Dict):
        """Validate a single client row."""
        # Required fields
        for field in ['full_legal_name', 'client_type', 'phone_primary']:
            if not row.get(field):
                self.errors.append({'row': row_num, 'field': field, 'error': f'{field} is required'})

        # Client type
        if row.get('client_type') and row['client_type'] not in ('INDIVIDUAL', 'SME', 'GROUP'):
            self.errors.append({'row': row_num, 'field': 'client_type', 'error': f'Invalid type: {row["client_type"]}. Must be INDIVIDUAL, SME, or GROUP'})

        # National ID format
        national_id = row.get('national_id_number', '')
        if national_id:
            result = self.engine.validate_national_id(str(national_id))
            if not result['valid']:
                self.errors.append({'row': row_num, 'field': 'national_id_number', 'error': result['message']})

        # Phone format
        phone = row.get('phone_primary', '')
        if phone:
            clean = str(phone).replace(' ', '').replace('-', '')
            if not re.match(r'^\+?\d{9,15}$', clean):
                self.warnings.append({'row': row_num, 'field': 'phone_primary', 'warning': 'Phone format may be invalid'})

        # Duplicate check
        if national_id:
            if Client.objects.filter(tenant=self.tenant, national_id_number=str(national_id)).exists():
                self.warnings.append({'row': row_num, 'field': 'national_id_number', 'warning': f'Duplicate: client with this ID already exists'})

        # Date of birth
        dob = row.get('date_of_birth')
        if dob:
            parsed = self._parse_date(dob)
            if parsed:
                age = (date.today() - parsed).days / 365.25
                if age < 18:
                    self.errors.append({'row': row_num, 'field': 'date_of_birth', 'error': 'Client must be at least 18 years old'})
            else:
                self.errors.append({'row': row_num, 'field': 'date_of_birth', 'error': 'Invalid date format. Use DD/MM/YYYY'})

        # Risk rating
        risk = row.get('risk_rating', '')
        if risk and risk not in ('LOW', 'MEDIUM', 'HIGH'):
            self.errors.append({'row': row_num, 'field': 'risk_rating', 'error': f'Invalid risk rating: {risk}'})

        # Income
        income = row.get('monthly_income')
        if income:
            try:
                val = Decimal(str(income))
                if val < 0:
                    self.errors.append({'row': row_num, 'field': 'monthly_income', 'error': 'Income cannot be negative'})
            except (InvalidOperation, ValueError):
                self.errors.append({'row': row_num, 'field': 'monthly_income', 'error': 'Invalid number'})

    # ─── Loan Validation ───

    def _validate_loan_row(self, row_num: int, row: Dict):
        """Validate a loan import row (for migrating existing portfolios)."""
        for field in ['client_number', 'principal_amount', 'outstanding_principal', 'interest_rate_pct', 'disbursement_date', 'status']:
            if not row.get(field):
                self.errors.append({'row': row_num, 'field': field, 'error': f'{field} is required'})

        # Client exists
        client_num = row.get('client_number', '')
        if client_num and not Client.objects.filter(tenant=self.tenant, client_number=str(client_num)).exists():
            self.errors.append({'row': row_num, 'field': 'client_number', 'error': f'Client {client_num} not found. Import clients first.'})

        # Amount validation
        for field in ['principal_amount', 'outstanding_principal']:
            val = row.get(field)
            if val:
                try:
                    d = Decimal(str(val))
                    if d < 0:
                        self.errors.append({'row': row_num, 'field': field, 'error': 'Amount cannot be negative'})
                except (InvalidOperation, ValueError):
                    self.errors.append({'row': row_num, 'field': field, 'error': 'Invalid number'})

        # Outstanding cannot exceed principal
        try:
            principal = Decimal(str(row.get('principal_amount', 0)))
            outstanding = Decimal(str(row.get('outstanding_principal', 0)))
            if outstanding > principal:
                self.errors.append({'row': row_num, 'field': 'outstanding_principal', 'error': 'Outstanding cannot exceed principal'})
        except (InvalidOperation, ValueError):
            pass

        # Status
        status = row.get('status', '')
        if status and status not in ('ACTIVE', 'DISBURSED', 'CLOSED', 'WRITTEN_OFF'):
            self.errors.append({'row': row_num, 'field': 'status', 'error': f'Invalid status: {status}'})

    # ─── Chart of Accounts Validation ───

    def _validate_coa_row(self, row_num: int, row: Dict):
        """Validate a chart of accounts row."""
        for field in ['account_code', 'account_name', 'account_type', 'normal_balance']:
            if not row.get(field):
                self.errors.append({'row': row_num, 'field': field, 'error': f'{field} is required'})

        acct_type = row.get('account_type', '')
        if acct_type and acct_type not in ('ASSET', 'LIABILITY', 'EQUITY', 'INCOME', 'EXPENSE'):
            self.errors.append({'row': row_num, 'field': 'account_type', 'error': f'Invalid type: {acct_type}'})

        balance = row.get('normal_balance', '')
        if balance and balance not in ('D', 'C'):
            self.errors.append({'row': row_num, 'field': 'normal_balance', 'error': 'Must be D (debit) or C (credit)'})

    # ─── Helpers ───

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        if isinstance(value, (datetime, date)):
            return value if isinstance(value, date) else value.date()
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except (ValueError, TypeError):
                continue
        return None


# ─── Validator Registry ───

VALIDATORS = {
    'CLIENTS': {
        'required_columns': ['full_legal_name', 'client_type', 'phone_primary'],
        'validate_row': ImportValidator._validate_client_row,
    },
    'LOANS': {
        'required_columns': ['client_number', 'principal_amount', 'outstanding_principal', 'interest_rate_pct', 'disbursement_date', 'status'],
        'validate_row': ImportValidator._validate_loan_row,
    },
    'CHART_OF_ACCOUNTS': {
        'required_columns': ['account_code', 'account_name', 'account_type', 'normal_balance'],
        'validate_row': ImportValidator._validate_coa_row,
    },
}


def run_import_validation(import_job_id: str) -> Dict:
    """Celery-compatible entry point for validating an import job."""
    job = ImportJob.objects.select_related('tenant').get(id=import_job_id)
    job.status = 'VALIDATING'
    job.save(update_fields=['status'])

    try:
        # Download file from Supabase Storage
        from supabase import create_client
        from django.conf import settings
        sb = create_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_ROLE_KEY)
        file_bytes = sb.storage.from_('imports').download(job.file_path)

        # Parse and validate
        validator = ImportValidator(str(job.tenant_id), job.import_type)
        rows = validator.parse_file(file_bytes, job.file_name)
        result = validator.validate(rows)

        # Update job
        job.status = 'VALIDATION_COMPLETE'
        job.total_rows = result['total_rows']
        job.valid_rows = result['valid_rows']
        job.error_rows = result['error_rows']
        job.warning_rows = result['warning_rows']
        job.validation_errors = result['errors']
        job.validation_warnings = result['warnings']
        job.save()

        return result

    except Exception as e:
        job.status = 'FAILED'
        job.error_message = str(e)[:500]
        job.save()
        logger.error(f'Import validation failed for {import_job_id}: {e}')
        raise
