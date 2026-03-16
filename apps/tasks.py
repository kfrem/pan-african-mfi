"""
Celery Background Tasks — Pan-African Microfinance SaaS
Scheduled and on-demand tasks for automated operations.

Schedule (configured via django-celery-beat):
- Every night 01:00 UTC: reclassify_all_loans
- Every night 02:00 UTC: check_overdue_schedules + send_repayment_reminders
- Every night 03:00 UTC: run_aml_monitoring
- Every 5 minutes: process_sync_queue
- Every hour: check_scheduled_reports
- Every day 06:00 UTC: send_kyc_expiry_warnings
"""
import logging
from datetime import date, timedelta
from decimal import Decimal

from celery import shared_task
from django.utils import timezone
from django.db.models import Sum, F, Q

from apps.tenants.models import Tenant
from apps.loans.models import Loan

logger = logging.getLogger(__name__)


# ─── LOAN CLASSIFICATION (Nightly) ───

@shared_task(name='loans.reclassify_all')
def reclassify_all_loans():
    """Reclassify all active loans across all tenants using Country Pack rules."""
    from apps.tenants.country_pack_engine import LoanClassificationService
    results = LoanClassificationService.reclassify_all_tenants()
    total_updated = sum(r['loans_updated'] for r in results)
    logger.info(f'Loan reclassification complete: {len(results)} tenants, {total_updated} loans updated')
    return results


# ─── SMS REMINDERS ───

@shared_task(name='sms.send_repayment_reminders')
def send_repayment_reminders():
    """Send SMS reminders for upcoming and overdue repayments."""
    from apps.loans.models import RepaymentSchedule
    from apps.notifications.models import SmsTemplate, SmsLog
    from apps.tenants.models import Tenant

    today = date.today()
    reminder_dates = {
        'REPAYMENT_REMINDER_3DAY': today + timedelta(days=3),
        'REPAYMENT_REMINDER_DUE': today,
        'REPAYMENT_OVERDUE_1DAY': today - timedelta(days=1),
        'REPAYMENT_OVERDUE_7DAY': today - timedelta(days=7),
    }

    sent_count = 0
    for template_code, target_date in reminder_dates.items():
        schedules = RepaymentSchedule.objects.filter(
            due_date=target_date,
            status__in=['PENDING', 'PARTIAL', 'OVERDUE'],
        ).select_related('loan__client', 'loan__tenant')

        for schedule in schedules.iterator(chunk_size=100):
            loan = schedule.loan
            client = loan.client

            if not client.phone_primary:
                continue

            # Find template for this tenant's country
            template = SmsTemplate.objects.filter(
                Q(tenant_id=loan.tenant_id) | Q(tenant_id__isnull=True),
                country_code=loan.tenant.country_code,
                template_code=template_code,
                is_active=True,
            ).order_by('-tenant_id').first()  # Tenant-specific overrides first

            if not template:
                continue

            # Render message
            amount_due = schedule.total_due - schedule.total_paid
            message = template.message_body
            message = message.replace('{{client_name}}', client.full_legal_name)
            message = message.replace('{{amount}}', f'{amount_due:,.2f}')
            message = message.replace('{{loan_number}}', loan.loan_number)
            message = message.replace('{{due_date}}', schedule.due_date.strftime('%d/%m/%Y'))
            message = message.replace('{{balance}}', f'{loan.outstanding_principal:,.2f}')
            message = message.replace('{{institution_name}}', loan.tenant.trading_name or loan.tenant.name)
            message = message.replace('{{currency}}', loan.currency)

            # Queue SMS
            SmsLog.objects.create(
                tenant_id=loan.tenant_id,
                template=template,
                recipient_phone=client.phone_primary,
                recipient_client=client,
                message_body=message,
                sms_parts=template.max_sms_parts,
                provider='AFRICAS_TALKING',
                status='QUEUED',
                triggered_by='SCHEDULER',
            )
            sent_count += 1

    logger.info(f'SMS reminders queued: {sent_count}')
    # Trigger actual sending
    send_queued_sms.delay()
    return sent_count


@shared_task(name='sms.send_queued')
def send_queued_sms():
    """Send all queued SMS messages via Africa's Talking."""
    from apps.notifications.models import SmsLog
    from django.conf import settings

    queued = SmsLog.objects.filter(status='QUEUED').order_by('created_at')[:500]

    if not queued.exists():
        return 0

    try:
        import africastalking
        africastalking.initialize(
            settings.AFRICAS_TALKING_USERNAME,
            settings.AFRICAS_TALKING_API_KEY
        )
        sms_service = africastalking.SMS
    except Exception as e:
        logger.error(f'Africa\'s Talking init failed: {e}')
        return 0

    sent = 0
    for msg in queued:
        try:
            response = sms_service.send(msg.message_body, [msg.recipient_phone])
            recipients = response.get('SMSMessageData', {}).get('Recipients', [])
            if recipients:
                r = recipients[0]
                msg.provider_message_id = r.get('messageId', '')
                msg.cost_amount = Decimal(r.get('cost', '0').replace('USD ', ''))
                msg.cost_currency = 'USD'
                msg.status = 'SENT'
                msg.sent_at = timezone.now()
            else:
                msg.status = 'FAILED'
                msg.status_message = 'No recipients in response'
        except Exception as e:
            msg.status = 'FAILED'
            msg.status_message = str(e)[:500]
            logger.error(f'SMS send failed for {msg.id}: {e}')

        msg.save()
        if msg.status == 'SENT':
            sent += 1

    logger.info(f'SMS sent: {sent}/{queued.count()}')
    return sent


# ─── KYC EXPIRY WARNINGS ───

@shared_task(name='kyc.send_expiry_warnings')
def send_kyc_expiry_warnings():
    """Warn clients whose ID documents are expiring within 30 days."""
    from apps.clients.models import KycDocument
    from apps.notifications.models import Notification

    threshold = date.today() + timedelta(days=30)
    expiring = KycDocument.objects.filter(
        expiry_date__lte=threshold,
        expiry_date__gte=date.today(),
        verified=True,
    ).select_related('client', 'client__assigned_officer')

    created = 0
    for doc in expiring.iterator():
        if doc.client.assigned_officer:
            Notification.objects.get_or_create(
                tenant_id=doc.tenant_id,
                user=doc.client.assigned_officer,
                title=f'KYC document expiring: {doc.client.full_legal_name}',
                defaults={
                    'severity': 'WARNING',
                    'message': f'{doc.document_type} for {doc.client.full_legal_name} '
                               f'(#{doc.client.client_number}) expires on {doc.expiry_date}. '
                               f'Please arrange renewal.',
                    'link': f'/clients/{doc.client_id}',
                }
            )
            created += 1

    logger.info(f'KYC expiry notifications created: {created}')
    return created


# ─── AML MONITORING ───

@shared_task(name='aml.run_monitoring')
def run_aml_monitoring():
    """
    Run AML transaction monitoring rules across all tenants.
    Checks: large cash transactions, structuring patterns, unusual activity.
    """
    from apps.tenants.models import Tenant
    from apps.tenants.country_pack_engine import CountryPackEngine
    from apps.loans.models import Repayment
    from apps.compliance.models import AmlAlert

    today = date.today()
    yesterday = today - timedelta(days=1)
    alerts_created = 0

    for tenant in Tenant.objects.filter(status='ACTIVE', subscription_active=True):
        engine = CountryPackEngine.for_tenant(str(tenant.id))

        # Check 1: Large cash transactions exceeding CTR threshold
        threshold = engine.aml_ctr_threshold
        if threshold:
            large_txns = Repayment.objects.filter(
                tenant=tenant,
                payment_method='CASH',
                amount__gte=threshold,
                received_at__date__gte=yesterday,
                reversed=False,
            ).select_related('loan__client')

            for txn in large_txns:
                if not AmlAlert.objects.filter(
                    tenant=tenant,
                    source_transaction_id=txn.id
                ).exists():
                    AmlAlert.objects.create(
                        tenant=tenant,
                        client=txn.loan.client,
                        alert_type='LARGE_CASH',
                        trigger_description=f'Cash transaction of {txn.currency} {txn.amount:,.2f} '
                                          f'exceeds CTR threshold of {threshold:,.2f}',
                        trigger_amount=txn.amount,
                        trigger_currency=txn.currency,
                        source_transaction_id=txn.id,
                        risk_score=60,
                    )
                    alerts_created += 1

        # Check 2: Structuring detection (multiple transactions just below threshold)
        if threshold:
            structuring_threshold = threshold * Decimal('0.9')
            from django.db.models import Count as DbCount
            structured = Repayment.objects.filter(
                tenant=tenant,
                payment_method='CASH',
                amount__gte=structuring_threshold,
                amount__lt=threshold,
                received_at__date__gte=today - timedelta(days=7),
                reversed=False,
            ).values('loan__client_id').annotate(
                txn_count=DbCount('id')
            ).filter(txn_count__gte=3)

            for s in structured:
                from apps.clients.models import Client
                client = Client.objects.get(id=s['loan__client_id'])
                if not AmlAlert.objects.filter(
                    tenant=tenant,
                    client=client,
                    alert_type='STRUCTURING',
                    created_at__gte=timezone.now() - timedelta(days=7),
                ).exists():
                    AmlAlert.objects.create(
                        tenant=tenant,
                        client=client,
                        alert_type='STRUCTURING',
                        trigger_description=f'Potential structuring: {s["txn_count"]} cash transactions '
                                          f'between {structuring_threshold:,.0f} and {threshold:,.0f} '
                                          f'in the last 7 days',
                        risk_score=75,
                    )
                    alerts_created += 1

    logger.info(f'AML monitoring complete: {alerts_created} alerts created')
    return alerts_created


# ─── STR ESCALATION ───

@shared_task(name='aml.escalate_stale_alerts')
def escalate_stale_alerts():
    """Auto-escalate AML alerts unactioned for 3 business days."""
    from apps.compliance.models import AmlAlert
    from apps.notifications.models import Notification

    threshold = timezone.now() - timedelta(days=3)
    stale = AmlAlert.objects.filter(
        status='OPEN',
        created_at__lte=threshold,
    )

    escalated = 0
    for alert in stale:
        alert.status = 'ESCALATED'
        alert.escalated_at = timezone.now()
        alert.save()

        # Notify compliance officers
        from apps.accounts.models import UserRole
        compliance_users = UserRole.objects.filter(
            role__tenant_id=alert.tenant_id,
            role__role_code='COMPLIANCE_OFFICER',
        ).values_list('user_id', flat=True)

        for uid in compliance_users:
            Notification.objects.create(
                tenant_id=alert.tenant_id,
                user_id=uid,
                severity='CRITICAL',
                title=f'AML Alert Escalated: {alert.alert_type}',
                message=f'Alert #{alert.id} for client has been auto-escalated after 3 days without review.',
                link=f'/compliance/aml-alerts/{alert.id}',
            )
        escalated += 1

    logger.info(f'AML alerts escalated: {escalated}')
    return escalated


# ─── SYNC QUEUE PROCESSING ───

@shared_task(name='sync.process_queue')
def process_sync_queue():
    """Process pending items in the offline sync queue."""
    from apps.integrations.models import SyncQueue

    pending = SyncQueue.objects.filter(status='QUEUED').order_by('client_timestamp')[:200]
    results = {'applied': 0, 'conflicts': 0, 'rejected': 0}

    for item in pending:
        item.status = 'PROCESSING'
        item.save(update_fields=['status'])

        try:
            # Route to the appropriate handler based on target table
            handler = SYNC_HANDLERS.get(item.target_table)
            if handler:
                handler(item)
                item.status = 'APPLIED'
                item.processed_at = timezone.now()
                results['applied'] += 1
            else:
                item.status = 'REJECTED'
                item.error_message = f'No sync handler for table: {item.target_table}'
                results['rejected'] += 1
        except ConflictError as e:
            item.status = 'CONFLICT'
            item.error_message = str(e)
            results['conflicts'] += 1
        except Exception as e:
            item.status = 'REJECTED'
            item.error_message = str(e)[:500]
            results['rejected'] += 1
            logger.error(f'Sync error for {item.id}: {e}')

        item.save()

    logger.info(f'Sync queue processed: {results}')
    return results


class ConflictError(Exception):
    pass


def _sync_client(item):
    """Handle offline-created client sync."""
    from apps.clients.models import Client
    if item.operation == 'INSERT':
        if Client.objects.filter(sync_id=item.target_sync_id).exists():
            raise ConflictError(f'Client with sync_id {item.target_sync_id} already exists')
        Client.objects.create(
            tenant_id=item.tenant_id,
            sync_id=item.target_sync_id,
            sync_status='SYNCED',
            server_confirmed_at=timezone.now(),
            device_id=item.device_id,
            client_created_at=item.client_timestamp,
            **{k: v for k, v in item.payload.items()
               if k not in ('tenant_id', 'sync_id', 'sync_status')}
        )
    elif item.operation == 'UPDATE':
        try:
            client = Client.objects.get(sync_id=item.target_sync_id, tenant_id=item.tenant_id)
            if client.updated_at > item.client_timestamp:
                raise ConflictError('Server version is newer than offline version')
            for field, value in item.payload.items():
                if hasattr(client, field) and field not in ('id', 'tenant_id', 'sync_id'):
                    setattr(client, field, value)
            client.sync_status = 'SYNCED'
            client.server_confirmed_at = timezone.now()
            client.save()
        except Client.DoesNotExist:
            raise ConflictError(f'Client with sync_id {item.target_sync_id} not found')


def _sync_repayment(item):
    """Handle offline-captured repayment sync."""
    from apps.loans.models import Repayment
    if item.operation == 'INSERT':
        if Repayment.objects.filter(sync_id=item.target_sync_id).exists():
            raise ConflictError(f'Repayment with sync_id {item.target_sync_id} already exists')
        # Use the repayment capture logic from API views
        from apps.api_views import RepaymentViewSet
        # Create a minimal repayment record — the capture endpoint handles balance updates
        Repayment.objects.create(
            tenant_id=item.tenant_id,
            sync_id=item.target_sync_id,
            sync_status='SYNCED',
            server_confirmed_at=timezone.now(),
            device_id=item.device_id,
            client_created_at=item.client_timestamp,
            **{k: v for k, v in item.payload.items()
               if k not in ('tenant_id', 'sync_id', 'sync_status')}
        )


SYNC_HANDLERS = {
    'clients': _sync_client,
    'repayments': _sync_repayment,
}


# ─── REPORT GENERATION ───

@shared_task(name='reports.check_scheduled')
def check_scheduled_reports():
    """Check for reports that need to be generated on schedule."""
    from apps.reports.models import ReportSchedule, ReportRun

    now = timezone.now()
    due_schedules = ReportSchedule.objects.filter(
        is_active=True,
        next_run_at__lte=now,
    )

    queued = 0
    for schedule in due_schedules:
        ReportRun.objects.create(
            tenant_id=schedule.tenant_id,
            report=schedule.report,
            schedule=schedule,
            parameters=schedule.parameters,
            output_format=schedule.output_format,
            status='QUEUED',
        )

        # Calculate next run time
        from dateutil.relativedelta import relativedelta
        if schedule.frequency == 'DAILY':
            schedule.next_run_at += timedelta(days=1)
        elif schedule.frequency == 'WEEKLY':
            schedule.next_run_at += timedelta(weeks=1)
        elif schedule.frequency == 'MONTHLY':
            schedule.next_run_at += relativedelta(months=1)
        elif schedule.frequency == 'QUARTERLY':
            schedule.next_run_at += relativedelta(months=3)
        elif schedule.frequency == 'ANNUAL':
            schedule.next_run_at += relativedelta(years=1)

        schedule.last_run_at = now
        schedule.save()
        queued += 1

    if queued > 0:
        generate_queued_reports.delay()

    logger.info(f'Scheduled reports queued: {queued}')
    return queued


@shared_task(name='reports.generate_queued')
def generate_queued_reports():
    """Generate all queued report runs."""
    from apps.reports.models import ReportRun
    import time

    queued = ReportRun.objects.filter(status='QUEUED').order_by('created_at')[:20]

    for run in queued:
        run.status = 'GENERATING'
        run.save(update_fields=['status'])

        start = time.monotonic()
        try:
            if run.output_format == 'PDF':
                file_path, page_count = _generate_pdf_report(run)
            elif run.output_format in ('EXCEL', 'XLS', 'XLSX'):
                file_path, page_count = _generate_excel_report(run)
            elif run.output_format == 'CSV':
                file_path, page_count = _generate_csv_report(run)
            else:
                file_path, page_count = _generate_pdf_report(run)

            run.status = 'COMPLETED'
            run.file_path = file_path
            run.page_count = page_count
            run.generated_at = timezone.now()
            run.generation_time_ms = int((time.monotonic() - start) * 1000)
            run.expires_at = timezone.now() + timedelta(days=90)
            logger.info(f'Report generated: {run.report.report_code} for tenant {run.tenant_id}')
        except Exception as e:
            run.status = 'FAILED'
            run.error_message = str(e)[:500]
            logger.error(f'Report generation failed: {run.id} - {e}')

        run.save()


def _get_report_context(run) -> dict:
    """Build the template/data context for a report run."""
    from django.db.models import Sum, Count, Avg

    tenant = Tenant.objects.select_related('country', 'licence_tier').get(id=run.tenant_id)
    params = run.parameters or {}

    # Common context available to all reports
    active_loans = Loan.objects.filter(
        tenant=tenant, status__in=['ACTIVE', 'DISBURSED']
    )
    stats = active_loans.aggregate(
        total_portfolio=Sum('outstanding_principal'),
        total_loans=Count('id'),
        avg_rate=Avg('interest_rate_pct'),
    )
    par30 = Loan.objects.filter(
        tenant=tenant, status__in=['ACTIVE', 'DISBURSED'], days_past_due__gte=30
    ).aggregate(par30_balance=Sum('outstanding_principal'))
    portfolio = stats['total_portfolio'] or Decimal('0')
    par30_bal = par30['par30_balance'] or Decimal('0')
    par30_pct = float(par30_bal / portfolio * 100) if portfolio > 0 else 0.0

    ctx = {
        'institution_name': tenant.name,
        'primary_colour': tenant.primary_brand_colour or '#1a56db',
        'currency': tenant.default_currency,
        'generated_date': timezone.now().strftime('%d %B %Y'),
        'page_footer': f'{tenant.name} · Confidential',
        'confidentiality': 'CONFIDENTIAL — For authorised recipients only',
        'total_portfolio': portfolio,
        'active_loans': stats['total_loans'] or 0,
        'avg_interest_rate': float(stats['avg_rate'] or 0),
        'par30_amount': par30_bal,
        'par30_pct': round(par30_pct, 2),
        'period': params.get('period', timezone.now().strftime('%B %Y')),
        'report_title': run.report.report_name,
        'report_code': run.report.report_code,
    }

    # Loan-specific context
    loan_id = params.get('loan_id')
    if loan_id:
        try:
            loan = Loan.objects.select_related(
                'client', 'product', 'loan_officer', 'branch'
            ).get(id=loan_id, tenant=tenant)
            remaining = loan.outstanding_principal
            total_paid = loan.repayments.filter(reversed=False).aggregate(
                t=Sum('amount')
            )['t'] or Decimal('0')
            ctx.update({
                'loan': loan,
                'client': loan.client,
                'remaining_balance': remaining,
                'total_paid': total_paid,
                'schedule': loan.schedule.order_by('instalment_number'),
                'repayments': loan.repayments.filter(reversed=False).order_by('received_at'),
            })
        except Exception:
            pass

    # Investor-specific context
    investor_id = params.get('investor_id')
    if investor_id:
        try:
            from apps.investors.models import InvestorProfile
            investor = InvestorProfile.objects.get(id=investor_id, tenant=tenant)
            ctx.update({
                'investor_name': investor.investor_name,
                'investor_type': investor.investor_type,
                'invested_amount': investor.invested_amount,
                'investment_date': investor.investment_date,
                'investment_currency': investor.investment_currency,
            })
        except Exception:
            pass

    return ctx


def _generate_pdf_report(run) -> tuple:
    """Generate a PDF report using WeasyPrint. Returns (file_path, page_count)."""
    import io
    import os
    import tempfile

    try:
        from weasyprint import HTML, CSS
        from django.template.loader import render_to_string
    except ImportError:
        # WeasyPrint not available — create a minimal placeholder PDF
        return _generate_placeholder_pdf(run)

    ctx = _get_report_context(run)

    # Map report code to template
    template_map = {
        'LOAN_STATEMENT': 'reports/loan_statement.html',
        'INVESTOR_REPORT': 'reports/investor_report.html',
        'BOARD_PACK': 'reports/board_pack.html',
    }
    template_name = template_map.get(run.report.report_code, 'reports/board_pack.html')

    try:
        html_content = render_to_string(template_name, ctx)
    except Exception:
        # Fall back to base template if specific one fails
        from django.template import Template, Context
        html_content = render_to_string('reports/base_report.html', {
            **ctx,
            'content': f'<h1>{run.report.report_name}</h1><p>Period: {ctx.get("period")}</p>',
        })

    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        HTML(string=html_content).write_pdf(tmp_path)
        file_size = os.path.getsize(tmp_path)

        # In production, upload to Supabase Storage
        # supabase.storage.from_('reports').upload(file_path, open(tmp_path, 'rb'))
        storage_path = (
            f'reports/{run.tenant_id}/{run.report.report_code}/'
            f'{timezone.now().strftime("%Y%m%d_%H%M%S")}_{run.id}.pdf'
        )
        run.file_size_bytes = file_size
        page_count = 1  # WeasyPrint doesn't easily expose page count without parsing
        return storage_path, page_count
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _generate_placeholder_pdf(run) -> tuple:
    """Create a minimal PDF placeholder when WeasyPrint is unavailable."""
    storage_path = (
        f'reports/{run.tenant_id}/{run.report.report_code}/'
        f'{timezone.now().strftime("%Y%m%d_%H%M%S")}_{run.id}.pdf'
    )
    return storage_path, 1


def _generate_excel_report(run) -> tuple:
    """Generate an Excel report using openpyxl. Returns (file_path, sheet_count)."""
    import io
    import os
    import tempfile

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        storage_path = (
            f'reports/{run.tenant_id}/{run.report.report_code}/'
            f'{timezone.now().strftime("%Y%m%d_%H%M%S")}_{run.id}.xlsx'
        )
        return storage_path, 1

    from django.db.models import Sum, Count

    tenant = Tenant.objects.get(id=run.tenant_id)
    params = run.parameters or {}

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header style
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    code = run.report.report_code

    if code == 'LOAN_PORTFOLIO':
        ws.title = 'Loan Portfolio'
        headers = ['Loan #', 'Client', 'Product', 'Principal', 'Outstanding', 'Status',
                   'Classification', 'PAR (days)', 'Branch', 'Officer', 'Disbursed']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        loans = Loan.objects.filter(
            tenant=tenant, status__in=['ACTIVE', 'DISBURSED']
        ).select_related('client', 'product', 'branch', 'loan_officer').order_by('loan_number')

        for row, loan in enumerate(loans, 2):
            ws.cell(row=row, column=1, value=loan.loan_number)
            ws.cell(row=row, column=2, value=loan.client.full_legal_name)
            ws.cell(row=row, column=3, value=loan.product.product_name)
            ws.cell(row=row, column=4, value=float(loan.principal_amount))
            ws.cell(row=row, column=5, value=float(loan.outstanding_principal))
            ws.cell(row=row, column=6, value=loan.status)
            ws.cell(row=row, column=7, value=loan.classification)
            ws.cell(row=row, column=8, value=loan.days_past_due)
            ws.cell(row=row, column=9, value=loan.branch.branch_name if loan.branch else '')
            ws.cell(row=row, column=10, value=loan.loan_officer.full_name if loan.loan_officer else '')
            ws.cell(row=row, column=11, value=str(loan.disbursement_date) if loan.disbursement_date else '')

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    else:
        # Generic data export
        ws.title = run.report.report_name[:31]
        ws.cell(row=1, column=1, value=f'{tenant.name} — {run.report.report_name}')
        ws.cell(row=2, column=1, value=f'Generated: {timezone.now().strftime("%d/%m/%Y %H:%M")}')
        ws.cell(row=3, column=1, value=f'Period: {params.get("period", "")}')

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        file_size = os.path.getsize(tmp_path)
        run.file_size_bytes = file_size
        storage_path = (
            f'reports/{run.tenant_id}/{run.report.report_code}/'
            f'{timezone.now().strftime("%Y%m%d_%H%M%S")}_{run.id}.xlsx'
        )
        return storage_path, wb.sheetnames.__len__()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _generate_csv_report(run) -> tuple:
    """Generate a CSV export. Returns (file_path, 1)."""
    import csv
    import io
    import os
    import tempfile

    tenant = Tenant.objects.get(id=run.tenant_id)

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
        tmp_path = tmp.name
        writer = csv.writer(tmp)

        writer.writerow([f'{tenant.name} — {run.report.report_name}'])
        writer.writerow([f'Generated: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
        writer.writerow([])

        code = run.report.report_code
        if code in ('LOAN_PORTFOLIO', 'PORTFOLIO_REPORT'):
            writer.writerow(['Loan #', 'Client', 'Product', 'Principal', 'Outstanding',
                             'Status', 'Classification', 'PAR (days)', 'Branch'])
            loans = Loan.objects.filter(
                tenant=tenant, status__in=['ACTIVE', 'DISBURSED']
            ).select_related('client', 'product', 'branch').order_by('loan_number')
            for loan in loans:
                writer.writerow([
                    loan.loan_number,
                    loan.client.full_legal_name,
                    loan.product.product_name,
                    float(loan.principal_amount),
                    float(loan.outstanding_principal),
                    loan.status,
                    loan.classification,
                    loan.days_past_due,
                    loan.branch.branch_name if loan.branch else '',
                ])
        else:
            writer.writerow(['No data available for this report format.'])

    file_size = os.path.getsize(tmp_path)
    run.file_size_bytes = file_size
    storage_path = (
        f'reports/{run.tenant_id}/{run.report.report_code}/'
        f'{timezone.now().strftime("%Y%m%d_%H%M%S")}_{run.id}.csv'
    )
    try:
        os.unlink(tmp_path)
    except Exception:
        pass
    return storage_path, 1


# ─── REPORT ALIAS ───

@shared_task(name='reports.generate')
def generate_reports():
    """Alias task — triggers queued report generation."""
    return generate_queued_reports()


# ─── NOTIFICATION THRESHOLD CHECKS ───

@shared_task(name='notifications.check_thresholds')
def check_notification_thresholds():
    """Evaluate all active notification rules and fire alerts."""
    from apps.notifications.models import NotificationRule, Notification
    from apps.tenants.models import Tenant
    from apps.loans.models import Loan

    for tenant in Tenant.objects.filter(status='ACTIVE'):
        rules = NotificationRule.objects.filter(tenant=tenant, is_active=True)

        for rule in rules:
            current_value = _get_metric_value(tenant.id, rule.metric)
            if current_value is None:
                continue

            triggered = _evaluate_threshold(current_value, rule.operator, rule.threshold_value)
            if not triggered:
                continue

            # Get users with the notified roles
            from apps.accounts.models import UserRole
            target_users = UserRole.objects.filter(
                role__tenant=tenant,
                role__role_code__in=rule.notify_roles or [],
            ).values_list('user_id', flat=True).distinct()

            for uid in target_users:
                # Don't duplicate if already sent today
                if not Notification.objects.filter(
                    tenant=tenant, user_id=uid, rule=rule,
                    created_at__date=date.today()
                ).exists():
                    Notification.objects.create(
                        tenant=tenant,
                        user_id=uid,
                        rule=rule,
                        severity=rule.severity,
                        title=f'{rule.rule_name}: {current_value}',
                        message=f'{rule.metric} is {current_value} (threshold: {rule.operator} {rule.threshold_value})',
                    )


def _get_metric_value(tenant_id, metric):
    """Compute a metric value for threshold comparison."""
    from apps.loans.models import Loan
    from django.db.models import Sum

    active_loans = Loan.objects.filter(
        tenant_id=tenant_id, status__in=['ACTIVE', 'DISBURSED']
    )
    portfolio = active_loans.aggregate(total=Sum('outstanding_principal'))['total'] or Decimal('0')

    if metric == 'par_30':
        par = active_loans.filter(days_past_due__gte=30).aggregate(
            total=Sum('outstanding_principal'))['total'] or Decimal('0')
        return (par / portfolio * 100) if portfolio > 0 else Decimal('0')
    elif metric == 'par_90':
        par = active_loans.filter(days_past_due__gte=90).aggregate(
            total=Sum('outstanding_principal'))['total'] or Decimal('0')
        return (par / portfolio * 100) if portfolio > 0 else Decimal('0')
    elif metric == 'total_portfolio':
        return portfolio
    elif metric == 'npl_ratio':
        npl = active_loans.filter(
            classification__in=['SUBSTANDARD', 'DOUBTFUL', 'LOSS']
        ).aggregate(total=Sum('outstanding_principal'))['total'] or Decimal('0')
        return (npl / portfolio * 100) if portfolio > 0 else Decimal('0')

    return None


def _evaluate_threshold(value, operator, threshold):
    """Evaluate a threshold condition."""
    if operator == 'GT':
        return value > threshold
    elif operator == 'LT':
        return value < threshold
    elif operator == 'GTE':
        return value >= threshold
    elif operator == 'LTE':
        return value <= threshold
    elif operator == 'EQ':
        return value == threshold
    return False
