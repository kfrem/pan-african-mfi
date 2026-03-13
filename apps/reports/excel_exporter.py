"""
Excel/CSV Export Engine — Pan-African Microfinance SaaS
Generates branded, professional Excel exports and CSV downloads.
All exports include a metadata sheet with export details.
"""
import io
import csv
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from django.utils import timezone

from apps.tenants.models import Tenant


class ExcelExporter:
    """Generate branded Excel workbooks with data tables and metadata."""

    def __init__(self, tenant_id: str, exported_by: str = ''):
        self.tenant = Tenant.objects.get(id=tenant_id)
        self.exported_by = exported_by
        self.now = timezone.now()

        # Brand colours
        primary_hex = (self.tenant.primary_brand_colour or '#1b3a6b').lstrip('#')
        self.primary_fill = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type='solid')
        self.alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        self.header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        self.body_font = Font(name='Calibri', size=10)
        self.number_font = Font(name='Calibri', size=10)
        self.title_font = Font(name='Calibri', bold=True, size=14, color=primary_hex)
        self.thin_border = Border(
            bottom=Side(style='thin', color='E2E8F0'),
        )

    def create_workbook(self, title: str, data: List[Dict[str, Any]],
                        columns: List[Dict[str, str]],
                        sheet_name: str = 'Data') -> bytes:
        """
        Create a branded Excel workbook.

        Args:
            title: Report title
            data: List of row dicts
            columns: List of {"key": "field_name", "label": "Column Header", "type": "text|number|currency|date|percent"}
            sheet_name: Name of the data sheet
        """
        wb = openpyxl.Workbook()

        # Data sheet
        ws = wb.active
        ws.title = sheet_name
        self._write_data_sheet(ws, title, data, columns)

        # Metadata sheet
        meta = wb.create_sheet('Export Info')
        self._write_metadata_sheet(meta, title, len(data))

        # Return as bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _write_data_sheet(self, ws, title: str, data: List[Dict], columns: List[Dict]):
        """Write the main data sheet with branding and formatting."""
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        sub = ws.cell(row=2, column=1,
                      value=f'{self.tenant.trading_name or self.tenant.name} · {self.now.strftime("%d %B %Y")}')
        sub.font = Font(name='Calibri', size=9, color='94A3B8', italic=True)

        # Header row (row 4)
        header_row = 4
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_def['label'])
            cell.font = self.header_font
            cell.fill = self.primary_fill
            cell.alignment = Alignment(horizontal='center' if col_def.get('type') in ('number', 'currency', 'percent') else 'left',
                                       vertical='center')
            cell.border = self.thin_border

        ws.row_dimensions[header_row].height = 24

        # Data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, col_def in enumerate(columns, 1):
                value = row_data.get(col_def['key'], '')

                # Convert Decimal to float for Excel
                if isinstance(value, Decimal):
                    value = float(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.body_font
                cell.border = self.thin_border

                # Alternating row colours
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = self.alt_fill

                # Number formatting
                col_type = col_def.get('type', 'text')
                if col_type == 'currency':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col_type == 'number':
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                elif col_type == 'percent':
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)):
                        cell.value = value / 100  # Excel expects 0.xx for %
                elif col_type == 'date':
                    cell.number_format = 'DD/MM/YYYY'

        # Auto-fit column widths
        for col_idx, col_def in enumerate(columns, 1):
            max_length = max(
                len(str(col_def['label'])),
                max((len(str(row.get(col_def['key'], ''))) for row in data), default=0)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

        # Freeze panes (header row)
        ws.freeze_panes = f'A{header_row + 1}'

    def _write_metadata_sheet(self, ws, title: str, row_count: int):
        """Write the metadata sheet with export details."""
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

        meta_data = [
            ('Export Information', ''),
            ('Report Title', title),
            ('Institution', self.tenant.trading_name or self.tenant.name),
            ('Country', self.tenant.country.country_name),
            ('Currency', self.tenant.default_currency),
            ('Exported By', self.exported_by),
            ('Export Date', self.now.strftime('%d %B %Y')),
            ('Export Time', self.now.strftime('%H:%M:%S UTC')),
            ('Total Rows', str(row_count)),
            ('', ''),
            ('CONFIDENTIAL', 'This document contains proprietary data. Do not distribute without authorisation.'),
        ]

        for row_idx, (label, value) in enumerate(meta_data, 1):
            cell_a = ws.cell(row=row_idx, column=1, value=label)
            cell_b = ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_a.font = Font(name='Calibri', bold=True, size=14)
            elif label == 'CONFIDENTIAL':
                cell_a.font = Font(name='Calibri', bold=True, color='DC2626')
                cell_b.font = Font(name='Calibri', italic=True, color='94A3B8', size=9)
            else:
                cell_a.font = Font(name='Calibri', bold=True, color='475569', size=10)
                cell_b.font = Font(name='Calibri', size=10)


class CSVExporter:
    """Generate CSV downloads with UTF-8 BOM for Excel compatibility."""

    @staticmethod
    def export(data: List[Dict[str, Any]], columns: List[Dict[str, str]]) -> bytes:
        buffer = io.StringIO()
        # UTF-8 BOM for Excel compatibility
        buffer.write('\ufeff')

        writer = csv.writer(buffer)
        # Header
        writer.writerow([col['label'] for col in columns])
        # Data
        for row in data:
            writer.writerow([row.get(col['key'], '') for col in columns])

        return buffer.getvalue().encode('utf-8')


# ─── Predefined Export Configurations ───

LOAN_BOOK_COLUMNS = [
    {'key': 'loan_number', 'label': 'Loan Number', 'type': 'text'},
    {'key': 'client_name', 'label': 'Client Name', 'type': 'text'},
    {'key': 'client_number', 'label': 'Client Number', 'type': 'text'},
    {'key': 'product_name', 'label': 'Product', 'type': 'text'},
    {'key': 'branch_name', 'label': 'Branch', 'type': 'text'},
    {'key': 'principal_amount', 'label': 'Principal', 'type': 'currency'},
    {'key': 'outstanding_principal', 'label': 'Outstanding', 'type': 'currency'},
    {'key': 'interest_rate_pct', 'label': 'Rate (%)', 'type': 'number'},
    {'key': 'disbursement_date', 'label': 'Disbursed', 'type': 'date'},
    {'key': 'maturity_date', 'label': 'Maturity', 'type': 'date'},
    {'key': 'status', 'label': 'Status', 'type': 'text'},
    {'key': 'classification', 'label': 'Classification', 'type': 'text'},
    {'key': 'days_past_due', 'label': 'DPD', 'type': 'number'},
    {'key': 'provision_amount', 'label': 'Provision', 'type': 'currency'},
    {'key': 'officer_name', 'label': 'Loan Officer', 'type': 'text'},
    {'key': 'is_insider_loan', 'label': 'Insider', 'type': 'text'},
]

CLIENT_LIST_COLUMNS = [
    {'key': 'client_number', 'label': 'Client Number', 'type': 'text'},
    {'key': 'full_legal_name', 'label': 'Full Name', 'type': 'text'},
    {'key': 'client_type', 'label': 'Type', 'type': 'text'},
    {'key': 'national_id_number', 'label': 'National ID', 'type': 'text'},
    {'key': 'phone_primary', 'label': 'Phone', 'type': 'text'},
    {'key': 'branch_name', 'label': 'Branch', 'type': 'text'},
    {'key': 'kyc_status', 'label': 'KYC Status', 'type': 'text'},
    {'key': 'risk_rating', 'label': 'Risk Rating', 'type': 'text'},
    {'key': 'is_pep', 'label': 'PEP', 'type': 'text'},
    {'key': 'is_insider', 'label': 'Insider', 'type': 'text'},
]

TRIAL_BALANCE_COLUMNS = [
    {'key': 'account_code', 'label': 'Account Code', 'type': 'text'},
    {'key': 'account_name', 'label': 'Account Name', 'type': 'text'},
    {'key': 'account_type', 'label': 'Type', 'type': 'text'},
    {'key': 'debit_balance', 'label': 'Debit', 'type': 'currency'},
    {'key': 'credit_balance', 'label': 'Credit', 'type': 'currency'},
]
