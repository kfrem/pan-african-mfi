"""
Excel / CSV Export Engine — Pan-African Microfinance SaaS
Generates branded Excel files with institution colours, metadata sheet,
and proper numeric formatting. CSV exports for simple data dumps.
"""
import io
import csv
from datetime import date, datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from apps.tenants.models import Tenant


class ExcelExporter:
    """Generate branded Excel files with institution theming."""

    def __init__(self, tenant_id: str, exported_by: str = ''):
        self.tenant = Tenant.objects.get(id=tenant_id)
        self.exported_by = exported_by
        self.brand_colour = (self.tenant.primary_brand_colour or '#1B3A6B').lstrip('#')
        self.wb = Workbook()

    def create_data_sheet(self, title: str, headers: List[str], rows: List[List[Any]],
                          column_widths: Optional[List[int]] = None,
                          number_columns: Optional[List[int]] = None) -> None:
        """Create a formatted data sheet with headers and rows."""
        ws = self.wb.active if self.wb.active.title == 'Sheet' else self.wb.create_sheet()
        ws.title = title[:31]  # Excel limit

        # Header styling
        header_fill = PatternFill(start_color=self.brand_colour, end_color=self.brand_colour, fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        data_font = Font(name='Calibri', size=10)
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

                # Alternating row colour
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

                # Number formatting
                if number_columns and col_idx in number_columns:
                    if isinstance(value, (int, float, Decimal)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

        # Column widths
        if column_widths:
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
        else:
            # Auto-fit based on header length
            for i, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 4, 12)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows) + 1}'

    def add_metadata_sheet(self, report_name: str, filters: Dict[str, str] = None) -> None:
        """Add a metadata sheet with export info (required by spec)."""
        ws = self.wb.create_sheet(title='_Metadata')

        meta_items = [
            ('Report Name', report_name),
            ('Institution', self.tenant.trading_name or self.tenant.name),
            ('Country', self.tenant.country.country_name),
            ('Currency', self.tenant.default_currency),
            ('Exported By', self.exported_by),
            ('Exported At', datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')),
            ('Date Range', filters.get('date_range', 'All') if filters else 'All'),
        ]
        if filters:
            for key, value in filters.items():
                if key != 'date_range':
                    meta_items.append((key.replace('_', ' ').title(), str(value)))

        label_font = Font(name='Calibri', bold=True, size=10, color=self.brand_colour)
        value_font = Font(name='Calibri', size=10)

        for row, (label, value) in enumerate(meta_items, 1):
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value).font = value_font

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def to_bytes(self) -> bytes:
        """Return the workbook as bytes for download/storage."""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        return buffer.getvalue()


class CSVExporter:
    """Simple CSV export with header row."""

    @staticmethod
    def export(headers: List[str], rows: List[List[Any]]) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([str(v) if v is not None else '' for v in row])
        return buffer.getvalue().encode('utf-8')


# ─── Pre-built Export Functions ───

def export_loan_book(tenant_id: str, exported_by: str = '',
                     status_filter: str = None) -> bytes:
    """Export the full loan book to Excel."""
    from apps.loans.models import Loan

    loans = Loan.objects.filter(tenant_id=tenant_id).select_related(
        'client', 'product', 'loan_officer', 'branch'
    ).order_by('loan_number')

    if status_filter:
        loans = loans.filter(status=status_filter)

    headers = [
        'Loan Number', 'Client Name', 'Client Number', 'Product', 'Branch',
        'Officer', 'Principal', 'Outstanding', 'Interest Rate (%)',
        'Term (Months)', 'Frequency', 'Status', 'Classification',
        'Days Past Due', 'Arrears', 'Provision', 'Disbursement Date',
        'Maturity Date', 'Insider', 'Override'
    ]

    rows = []
    for loan in loans:
        rows.append([
            loan.loan_number,
            loan.client.full_legal_name,
            loan.client.client_number,
            loan.product.product_name,
            loan.branch.branch_name,
            loan.loan_officer.full_name,
            float(loan.principal_amount),
            float(loan.outstanding_principal),
            float(loan.interest_rate_pct),
            loan.term_months,
            loan.repayment_frequency,
            loan.status,
            loan.classification,
            loan.days_past_due,
            float(loan.arrears_amount),
            float(loan.provision_amount),
            str(loan.disbursement_date or ''),
            str(loan.maturity_date or ''),
            'Yes' if loan.is_insider_loan else 'No',
            'Yes' if loan.override_flag else 'No',
        ])

    exporter = ExcelExporter(tenant_id, exported_by)
    exporter.create_data_sheet(
        'Loan Book', headers, rows,
        column_widths=[15, 25, 15, 20, 15, 20, 15, 15, 12, 10, 12, 15, 15, 10, 15, 15, 15, 15, 8, 8],
        number_columns=[7, 8, 9, 15, 16]
    )
    exporter.add_metadata_sheet('Full Loan Book', {
        'status_filter': status_filter or 'All',
        'total_loans': str(len(rows)),
    })
    return exporter.to_bytes()


def export_client_list(tenant_id: str, exported_by: str = '') -> bytes:
    """Export client list to Excel."""
    from apps.clients.models import Client

    clients = Client.objects.filter(
        tenant_id=tenant_id, deleted_at__isnull=True
    ).select_related('branch', 'assigned_officer').order_by('client_number')

    headers = [
        'Client Number', 'Full Name', 'Type', 'Gender', 'Phone',
        'National ID', 'KYC Status', 'Risk Rating', 'Branch',
        'Officer', 'PEP', 'Insider', 'Created'
    ]

    rows = [[
        c.client_number, c.full_legal_name, c.client_type, c.gender, c.phone_primary,
        c.national_id_number, c.kyc_status, c.risk_rating,
        c.branch.branch_name if c.branch else '',
        c.assigned_officer.full_name if c.assigned_officer else '',
        'Yes' if c.is_pep else 'No',
        'Yes' if c.is_insider else 'No',
        str(c.created_at.date()),
    ] for c in clients]

    exporter = ExcelExporter(tenant_id, exported_by)
    exporter.create_data_sheet('Clients', headers, rows,
        column_widths=[15, 25, 12, 8, 18, 20, 12, 10, 15, 20, 6, 8, 12])
    exporter.add_metadata_sheet('Client List', {'total_clients': str(len(rows))})
    return exporter.to_bytes()


def export_trial_balance(tenant_id: str, period_id: str, exported_by: str = '') -> bytes:
    """Export trial balance for an accounting period."""
    from apps.ledger.models import GlAccount, GlEntry, AccountingPeriod
    from django.db.models import Sum

    period = AccountingPeriod.objects.get(id=period_id, tenant_id=tenant_id)

    accounts = GlAccount.objects.filter(
        tenant_id=tenant_id, is_active=True, is_header=False
    ).order_by('account_code')

    headers = ['Account Code', 'Account Name', 'Type', 'Debit', 'Credit', 'Balance']
    rows = []
    total_debit = Decimal('0')
    total_credit = Decimal('0')

    for account in accounts:
        debits = GlEntry.objects.filter(
            tenant_id=tenant_id, account=account,
            transaction__period=period
        ).aggregate(total=Sum('debit_amount'))['total'] or Decimal('0')

        credits = GlEntry.objects.filter(
            tenant_id=tenant_id, account=account,
            transaction__period=period
        ).aggregate(total=Sum('credit_amount'))['total'] or Decimal('0')

        balance = debits - credits if account.normal_balance == 'D' else credits - debits

        if debits > 0 or credits > 0:
            rows.append([
                account.account_code,
                account.account_name,
                account.account_type,
                float(debits),
                float(credits),
                float(balance),
            ])
            total_debit += debits
            total_credit += credits

    # Total row
    rows.append(['', 'TOTAL', '', float(total_debit), float(total_credit), float(total_debit - total_credit)])

    exporter = ExcelExporter(tenant_id, exported_by)
    exporter.create_data_sheet('Trial Balance', headers, rows,
        column_widths=[15, 30, 12, 18, 18, 18],
        number_columns=[4, 5, 6])
    exporter.add_metadata_sheet('Trial Balance', {
        'period': period.period_name,
        'date_range': f'{period.start_date} to {period.end_date}',
    })
    return exporter.to_bytes()
